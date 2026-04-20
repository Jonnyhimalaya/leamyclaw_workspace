#!/usr/bin/env python3
"""
ECL Excel Writer — produces Dano's template-matching workbook

Sheets:
  1. ECL Calc         — per-invoice detail (matches Dano's layout to the cell)
  2. Summary          — pivot by Subsidiary × Bucket
  3. Movement         — pivot by Comment × Subsidiary  
  4. Reconciliation   — audit trail block
  5. Flags            — rows needing human review (exclusions, anomalies)
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
BOLD  = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '#,##0.00_);[Red](#,##0.00);"–"'
CURRENCY_INT_FMT = '#,##0_);[Red](#,##0);"–"'
PERCENT_FMT = "0.00%"
DATE_FMT = "dd/mm/yyyy"


def _style_header_row(ws, row_idx: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, max_width: int = 40):
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            s = str(cell.value)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


# ---------------------------------------------------------------------------
# Write ECL Calc sheet — matches Dano's template layout exactly
# ---------------------------------------------------------------------------
COLS_IN_ORDER = [
    "Subsidiary", "Company Name", "Country Name", "Transaction Type",
    "Invoice Date", "Due Date", "Invoice No", "Memo", "Invoice Memo",
    "Age", "Invoice CCY", "Amount Due (Foreign Currency)",
    "Amount Due (Func Currency)", "Bucket Feb-26", "DSO", "Bucket", "Rate",
    "ECL Amount Mar-26", "ECL Amount Feb-26", "ECL Movement",
    "Date", "Comment", "Account Owner", "Entity: [IONO2C] Salesforce Account ID",
]


def _prepare_ecl_calc_frame(df: pd.DataFrame, period_label: str, period_prior_label: str) -> pd.DataFrame:
    """Transform the pipeline output into the 24-column template shape."""
    out = pd.DataFrame()

    out["Subsidiary"] = df["Subsidiary"]
    out["Company Name"] = df["Company Name"]
    out["Country Name"] = df.get("Country Name", "")
    out["Transaction Type"] = df["Transaction Type"]
    out["Invoice Date"] = df["_inv_date"]
    out["Due Date"] = df["_due_date"]
    out["Invoice No"] = df["Invoice No"]
    out["Memo"] = df["Memo"].fillna("")
    out["Invoice Memo"] = df["_key"]
    out["Age"] = df["Age"]
    out["Invoice CCY"] = df.get("Invoice CCY", "")
    out["Amount Due (Foreign Currency)"] = df["_amt_foreign"].round(2)
    out["Amount Due (Func Currency)"] = df["_amt_func"].round(2)
    out["Bucket Feb-26"] = df["_prior_bucket"].fillna("")
    out["DSO"] = df["DSO"]
    out["Bucket"] = df["Bucket"].fillna("")
    out["Rate"] = df["Rate"].round(4)
    out[f"ECL Amount {period_label}"] = df["ECL Amount"].round(0)
    out[f"ECL Amount {period_prior_label}"] = df["_prior_ecl"].round(0)
    out["ECL Movement"] = df["ECL Movement"].round(0)
    out["Date"] = period_label.split("-")[0]  # "Mar" from "Mar-26"
    out["Comment"] = df["Comment"].fillna("")
    out["Account Owner"] = df.get("Account Owner", "")
    out["Entity: [IONO2C] Salesforce Account ID"] = df.get("Salesforce ID", "")

    return out


# ---------------------------------------------------------------------------
# Main writer
# ---------------------------------------------------------------------------
def write_ecl_workbook(
    pipeline_output: pd.DataFrame,
    out_path: Path,
    period_end: datetime,
    source_file: str,
    source_hash: str,
    division: str,
    rates: dict,
    rate_table_version: str = "group-wide v2026.04 (KPMG-sourced)",
):
    period_label = period_end.strftime("%b-%y")  # e.g. "Mar-26"
    prior_dt = (period_end.replace(day=1) - pd.Timedelta(days=1))
    prior_label = prior_dt.strftime("%b-%y")     # e.g. "Feb-26"

    wb = Workbook()

    # ============ Sheet 1: ECL Calc ============
    ws = wb.active
    ws.title = "ECL Calc"

    df = _prepare_ecl_calc_frame(pipeline_output, period_label, prior_label)

    # Row 1: metadata header (period-end + grand totals)
    total_curr = pipeline_output["ECL Amount"].sum()
    total_prior = pipeline_output["_prior_ecl"].sum()
    total_move = pipeline_output["ECL Movement"].sum()

    ws.cell(row=1, column=15, value=period_end.strftime("%d/%m/%Y"))
    ws.cell(row=1, column=18, value=round(total_curr, 0)).number_format = CURRENCY_INT_FMT
    ws.cell(row=1, column=19, value=round(total_prior, 0)).number_format = CURRENCY_INT_FMT
    ws.cell(row=1, column=20, value=round(total_move, 0)).number_format = CURRENCY_INT_FMT
    for col in [15, 18, 19, 20]:
        ws.cell(row=1, column=col).font = BOLD

    # Row 2: column headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=2, column=col_idx, value=col_name)
    _style_header_row(ws, 2, len(df.columns))

    # Rows 3+: data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)

    # Formatting
    col_map = {name: i + 1 for i, name in enumerate(df.columns)}
    for row_idx in range(3, 3 + len(df)):
        # Dates
        for col_name in ["Invoice Date", "Due Date"]:
            c = ws.cell(row=row_idx, column=col_map[col_name])
            c.number_format = DATE_FMT
        # Currency
        for col_name in ["Amount Due (Foreign Currency)", "Amount Due (Func Currency)"]:
            c = ws.cell(row=row_idx, column=col_map[col_name])
            c.number_format = CURRENCY_FMT
        # Integer currency for ECL columns
        for col_name in [f"ECL Amount {period_label}", f"ECL Amount {prior_label}", "ECL Movement"]:
            c = ws.cell(row=row_idx, column=col_map[col_name])
            c.number_format = CURRENCY_INT_FMT
        # Rate as decimal 4dp (keeping Dano's format)
        c = ws.cell(row=row_idx, column=col_map["Rate"])
        c.number_format = "0.00"

    ws.freeze_panes = "A3"
    _autosize(ws, max_width=32)

    # ============ Sheet 2: Summary (pivot by Subsidiary × Bucket) ============
    ws2 = wb.create_sheet("Summary")
    active = pipeline_output.dropna(subset=["Bucket"]).copy()
    active["Bucket"] = active["Bucket"].astype(str)

    pivot = active.pivot_table(
        index="Subsidiary",
        columns="Bucket",
        values="ECL Amount",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="TOTAL",
    )
    # Ensure bucket columns are A..G order even if some missing
    all_buckets = ["A", "B", "C", "D", "E", "F", "G"]
    for b in all_buckets:
        if b not in pivot.columns:
            pivot[b] = 0
    pivot = pivot[[c for c in all_buckets if c in pivot.columns] + (["TOTAL"] if "TOTAL" in pivot.columns else [])]

    # Header
    ws2.cell(row=1, column=1, value=f"ECL Provision Summary — {division} — {period_label}")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws2.cell(row=3, column=1, value="Subsidiary")
    for col_idx, b in enumerate(pivot.columns, 2):
        ws2.cell(row=3, column=col_idx, value=f"Bucket {b}")
    _style_header_row(ws2, 3, len(pivot.columns) + 1)

    for r_idx, (sub, row) in enumerate(pivot.iterrows(), 4):
        ws2.cell(row=r_idx, column=1, value=sub)
        if sub == "TOTAL":
            ws2.cell(row=r_idx, column=1).font = BOLD
        for c_idx, b in enumerate(pivot.columns, 2):
            c = ws2.cell(row=r_idx, column=c_idx, value=round(row[b], 0))
            c.number_format = CURRENCY_INT_FMT
            if sub == "TOTAL":
                c.font = BOLD
                c.fill = SUBTOTAL_FILL

    ws2.freeze_panes = "B4"
    _autosize(ws2, max_width=45)

    # ============ Sheet 3: Movement Analysis ============
    ws3 = wb.create_sheet("Movement")
    ws3.cell(row=1, column=1, value=f"ECL Movement — {prior_label} → {period_label}").font = Font(bold=True, size=14)

    mv = pipeline_output.copy()
    mv["Comment"] = mv["Comment"].fillna("(unclassified)")
    mv_pivot = mv.pivot_table(
        index="Comment",
        columns="Subsidiary",
        values="ECL Movement",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="TOTAL",
    )

    ws3.cell(row=3, column=1, value="Comment")
    for col_idx, sub in enumerate(mv_pivot.columns, 2):
        ws3.cell(row=3, column=col_idx, value=sub)
    _style_header_row(ws3, 3, len(mv_pivot.columns) + 1)

    for r_idx, (cmt, row) in enumerate(mv_pivot.iterrows(), 4):
        ws3.cell(row=r_idx, column=1, value=cmt)
        if cmt == "TOTAL":
            ws3.cell(row=r_idx, column=1).font = BOLD
        for c_idx, sub in enumerate(mv_pivot.columns, 2):
            c = ws3.cell(row=r_idx, column=c_idx, value=round(row[sub], 0))
            c.number_format = CURRENCY_INT_FMT
            if cmt == "TOTAL":
                c.font = BOLD
                c.fill = SUBTOTAL_FILL

    ws3.freeze_panes = "B4"
    _autosize(ws3, max_width=45)

    # ============ Sheet 4: Reconciliation ============
    ws4 = wb.create_sheet("Reconciliation")
    ws4.cell(row=1, column=1, value="Audit Trail & Reconciliation").font = Font(bold=True, size=14)

    recon = [
        ("Run timestamp (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Source file",         source_file),
        ("Source SHA256",       source_hash),
        ("Period end",          period_end.strftime("%d/%m/%Y")),
        ("Division",            division),
        ("Rate table version",  rate_table_version),
        ("", ""),
        ("Row counts", ""),
        ("  Output rows",                   len(pipeline_output)),
        ("  Current-period active rows",    len(pipeline_output.dropna(subset=["Bucket"]))),
        ("  Prior-only (Paid/Credited)",    len(pipeline_output[pipeline_output["Bucket"].isna() | (pipeline_output["Bucket"]=="")])),
        ("", ""),
        ("Totals (USD)", ""),
        ("  Outstanding AR (current)",      round(pipeline_output["_amt_func"].fillna(0).sum(), 2)),
        (f"  ECL Provision {period_label}",  round(total_curr, 2)),
        (f"  ECL Provision {prior_label}",   round(total_prior, 2)),
        ("  ECL Movement",                  round(total_move, 2)),
        ("", ""),
        ("Rate table used", ""),
    ]
    for bucket, rate in sorted(rates.items()):
        recon.append((f"  Bucket {bucket}", f"{rate*100:.2f}%"))

    for r_idx, (label, val) in enumerate(recon, 3):
        ws4.cell(row=r_idx, column=1, value=label)
        ws4.cell(row=r_idx, column=2, value=val)
        if isinstance(val, (int, float)) and not isinstance(val, bool) and label.strip():
            ws4.cell(row=r_idx, column=2).number_format = CURRENCY_FMT if "$" in str(val) or "Total" in label or "Provision" in label or "Movement" in label or "Outstanding" in label else "0"
        if label and not label.startswith(" ") and ":" not in label and val == "":
            ws4.cell(row=r_idx, column=1).font = BOLD

    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 35

    # ============ Sheet 5: Flags ============
    ws5 = wb.create_sheet("Flags")
    ws5.cell(row=1, column=1, value="Rows Requiring Human Review").font = Font(bold=True, size=14)
    ws5.cell(row=2, column=1, value=(
        "Deliberate exclusions (Dealogic LLC), anomalies, and any rows "
        "flagged for manual inspection before sign-off."
    ))

    # Dealogic LLC exclusions
    excluded = pipeline_output[pipeline_output["Subsidiary"].astype(str).str.strip() == "Dealogic LLC"].copy()
    ws5.cell(row=4, column=1, value=f"Dealogic LLC exclusions: {len(excluded)} rows, "
                                      f"outstanding ${excluded['_amt_func'].sum():,.2f}").font = BOLD

    flag_cols = ["Subsidiary", "Company Name", "Invoice No", "DSO", "Bucket", "_amt_func"]
    ws5.cell(row=6, column=1, value="Subsidiary")
    ws5.cell(row=6, column=2, value="Company Name")
    ws5.cell(row=6, column=3, value="Invoice No")
    ws5.cell(row=6, column=4, value="DSO")
    ws5.cell(row=6, column=5, value="Bucket")
    ws5.cell(row=6, column=6, value="Outstanding (Func)")
    _style_header_row(ws5, 6, 6)

    for r_idx, (_, row) in enumerate(excluded.iterrows(), 7):
        ws5.cell(row=r_idx, column=1, value=row.get("Subsidiary"))
        ws5.cell(row=r_idx, column=2, value=row.get("Company Name"))
        ws5.cell(row=r_idx, column=3, value=row.get("Invoice No"))
        ws5.cell(row=r_idx, column=4, value=row.get("DSO"))
        ws5.cell(row=r_idx, column=5, value=row.get("Bucket"))
        c = ws5.cell(row=r_idx, column=6, value=row.get("_amt_func"))
        c.number_format = CURRENCY_FMT

    ws5.freeze_panes = "A7"
    _autosize(ws5, max_width=40)

    # Save
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# CLI entry — run after ecl_calculator.py
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from ecl_calculator import run, RATES_ANALYTICS_MARKETS, file_hash

    NS_PATH    = Path("/home/jonny/.openclaw/media/inbound/Netsuite_Data---7539c361-7a6b-4bb0-b989-2ee4ad983f0c.csv")
    PRIOR_PATH = Path("/home/jonny/.openclaw/media/inbound/ECL_March_2026---216e4ea9-f6bf-40cc-b45e-d9a15814bc75.csv")
    PERIOD_END = datetime(2026, 3, 31)

    print("\n" + "="*70)
    print("Running pipeline …")
    print("="*70)
    result = run(NS_PATH, PRIOR_PATH, PERIOD_END, division="ION Analytics")

    out = Path("/home/jonny/.openclaw/workspace/consultancy/clients/accountant/skill-build/Dano_ECL_Mar-26_replica.xlsx")
    print(f"\n{'='*70}")
    print(f"Writing Excel workbook → {out.name}")
    print("="*70)

    write_ecl_workbook(
        pipeline_output=result,
        out_path=out,
        period_end=PERIOD_END,
        source_file=NS_PATH.name,
        source_hash=file_hash(NS_PATH),
        division="ION Analytics",
        rates=RATES_ANALYTICS_MARKETS,
    )
    print(f"\n✅ Written: {out}")
    print(f"   Size: {out.stat().st_size / 1024:.1f} KB")
